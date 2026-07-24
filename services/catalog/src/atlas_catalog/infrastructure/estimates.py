"""Lightweight row/column estimates — not Phase 4 profiling."""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path

from atlas_catalog.domain import DatasetFormat


def estimate_shape(path: Path, fmt: DatasetFormat, encoding: str | None) -> tuple[int | None, int | None]:
    """Return (row_estimate, column_estimate) without full EDA."""
    try:
        if fmt in {DatasetFormat.CSV, DatasetFormat.TSV}:
            delim = "," if fmt is DatasetFormat.CSV else "\t"
            enc = encoding or "utf-8"
            with path.open("r", encoding=enc, errors="replace", newline="") as fh:
                reader = csv.reader(fh, delimiter=delim)
                header = next(reader, None)
                cols: int | None = len(header) if header else 0
                rows: int | None = sum(1 for _ in reader)
                return rows, cols
        if fmt is DatasetFormat.JSON:
            text = path.read_text(encoding=encoding or "utf-8", errors="replace")
            data = json.loads(text)
            if isinstance(data, list):
                cols = len(data[0].keys()) if data and isinstance(data[0], dict) else None
                return len(data), cols
            if isinstance(data, dict):
                return 1, len(data)
            return None, None
        if fmt is DatasetFormat.EXCEL:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = max((ws.max_row or 1) - 1, 0)
            cols = ws.max_column or 0
            wb.close()
            return rows, cols
        if fmt is DatasetFormat.PARQUET:
            import pyarrow.parquet as pq

            pf = pq.ParquetFile(path)
            meta = pf.metadata
            rows = int(meta.num_rows) if meta is not None else None
            cols = len(pf.schema_arrow.names)
            return rows, cols
        # ZIP or unrecognized — estimates deferred
        return None, None
    except Exception:  # noqa: BLE001 — estimates are best-effort
        return None, None


def estimate_shape_bytes(
    data: bytes, fmt: DatasetFormat, encoding: str | None
) -> tuple[int | None, int | None]:
    """Estimate from in-memory bytes (small files / tests)."""
    if fmt in {DatasetFormat.CSV, DatasetFormat.TSV, DatasetFormat.JSON}:
        text = data.decode(encoding or "utf-8", errors="replace")
        if fmt is DatasetFormat.JSON:
            try:
                parsed = json.loads(text)
            except json.JSONDecodeError:
                return None, None
            if isinstance(parsed, list):
                cols = len(parsed[0].keys()) if parsed and isinstance(parsed[0], dict) else None
                return len(parsed), cols
            if isinstance(parsed, dict):
                return 1, len(parsed)
            return None, None
        delim = "," if fmt is DatasetFormat.CSV else "\t"
        reader = csv.reader(io.StringIO(text), delimiter=delim)
        header = next(reader, None)
        cols = len(header) if header else 0
        rows = sum(1 for _ in reader)
        return rows, cols
    return None, None
